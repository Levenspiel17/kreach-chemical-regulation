"""
K-REACH 화학물질 엑셀 변환 스크립트
- 입력: 화학물질_YYYYMMDD.xlsx (13열, A~M)
- 출력: 화학물질_YYYYMMDD_변환.xlsx (20열, A~T)

추가 열 (N~T):
  N: 인체급성유해성 (%)  - K열 "유독물질 : 인체급성유해성 : X%" 에서 추출
  O: 인체만성유해성 (%)  - K열 "유독물질 : 인체만성유해성 : X%" 에서 추출
  P: 생태유해성 (%)      - K열 "유독물질 : 생태유해성 : X%" 에서 추출
  Q: 제한물질 (%)        - K열 "제한물질 : ... 이를 X% 이상" 에서 추출
  R: 금지물질 (%)        - K열 "금지물질 : ... 이를 X% 이상" 에서 추출
  S: 허가물질 (%)        - K열 "허가물질 : ... 이를 X% 이상" 에서 추출
  T: 사고대비물질        - G열 값 복사
"""

import re
import sys
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def parse_pct(text, keyword):
    """K열 텍스트에서 특정 키워드의 % 수치를 추출한다."""
    if not text:
        return None
    pattern = rf'{re.escape(keyword)}\s*:\s*([\d.]+)\s*%'
    m = re.search(pattern, text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def parse_regulated_pct(text, reg_type):
    """제한/금지/허가물질 함량 기준 % 추출 - '이를 X% 이상' 패턴"""
    if not text:
        return None
    # 해당 규제 유형 세그먼트 찾기
    segments = text.split(' / ')
    for seg in segments:
        if reg_type in seg:
            m = re.search(r'이를\s*([\d.]+)\s*%\s*이상', seg)
            if m:
                try:
                    return float(m.group(1))
                except ValueError:
                    pass
    return None


def transform(input_path, output_path):
    print(f"\n변환 시작: {input_path}")
    wb_in = openpyxl.load_workbook(input_path, data_only=True)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = ws_in.title

    # 헤더 작성
    original_headers = [cell.value for cell in ws_in[1]]
    new_headers = [
        '인체급성유해성', '인체만성유해성', '생태유해성',
        '제한물질', '금지물질', '허가물질', '사고대비물질'
    ]
    all_headers = original_headers + new_headers
    ws_out.append(all_headers)

    # 헤더 스타일 - 원본열: 회색, 추가열: 하늘색
    base_fill = PatternFill("solid", fgColor="D9D9D9")
    new_fill  = PatternFill("solid", fgColor="BDD7EE")
    bold = Font(bold=True)
    for i, cell in enumerate(ws_out[1]):
        cell.font = bold
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.fill = new_fill if i >= len(original_headers) else base_fill

    # 열 너비 설정
    col_widths = [8,15,45,20,12,30,10,18,12,10,55,12,10,
                  12,12,12,12,12,12,12]
    for i, w in enumerate(col_widths, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    # 데이터 변환
    total = ws_in.max_row - 1
    errors = []
    for row_idx, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=2):
        row = list(row)
        k_text = row[10] if len(row) > 10 else None  # K열 (index 10)
        g_val  = row[6]  if len(row) > 6  else None  # G열 (index 6)

        n = parse_pct(k_text, '인체급성유해성')
        o = parse_pct(k_text, '인체만성유해성')
        p = parse_pct(k_text, '생태유해성')
        q = parse_regulated_pct(k_text, '제한물질')
        r = parse_regulated_pct(k_text, '금지물질')
        s = parse_regulated_pct(k_text, '허가물질')
        t = g_val  # 사고대비물질 = G열 복사

        # K열 있는데 N~S 모두 None이면 파싱 실패 기록
        if k_text and all(v is None for v in [n, o, p, q, r, s]):
            errors.append(row_idx)

        ws_out.append(row + [n, o, p, q, r, s, t])

    wb_out.save(output_path)
    print(f"✅ 변환 완료: {output_path}")
    print(f"   총 {total:,}행 처리 / 파싱 실패 의심 행: {len(errors)}건")
    if errors[:5]:
        print(f"   실패 행 예시: {errors[:5]}")
    return total, len(errors)


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("사용법: python transform.py <입력.xlsx> <출력.xlsx>")
        sys.exit(1)
    transform(sys.argv[1], sys.argv[2])
